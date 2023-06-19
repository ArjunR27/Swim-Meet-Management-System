from psych_sheet import *
from openpyxl import Workbook

class HeatSheet:

    # Read from the psych_sheet text file, allowing for the ranking to be seen
    # From here create new txt file, with the heat number and lane number
    def __init__(self):
        self.dict_of_events = {'50YDButterfly': [], '100YDButterfly': [], '200YDButterfly': [],
                               '50YDBackstroke': [], '100YDBackstroke': [], '200YDBackstroke': [],
                               '50YDBreaststroke': [], '100YDBreaststroke': [], '200YDBreaststroke': [],
                               '50YDFreestyle': [], '100YDFreestyle': [], '200YDFreestyle': [],
                               '500YDFreestyle': [], '1000YDFreestyle': [], '1650YDFreestyle': [],
                               '100YDIM': [], '200YDIM': [], '400YDIM': []}

    def parser(self, input):
        global heat_sheet
        heat_sheet = HeatSheet()
        with open(input, 'r') as in_file:
            heat = 1
            lane_count = 0
            for line in in_file:
                # Once the lane count reaches 8, we need to increment to the next heat
                if lane_count > 7:
                    heat += 1
                    lane_count = 0
                lst = line.split()
                if len(lst) == 0:
                    heat = 1
                    lane_count = 0
                else:
                    event, f_name, l_name, team, seed_time = lst[0], lst[1], lst[2], lst[3], lst[4]
                    lane_order = [4, 5, 3, 6, 2, 7, 1, 8]
                    lane_num = lane_order[lane_count]
                    lane_count += 1
                    person = (event, f_name, l_name, team, seed_time, heat, lane_num)
                    p_event = person[0]
                    self.dict_of_events[p_event] += [person]

    def create_heat_sheet(self, output):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Heat Sheet"

        row = 1
        sheet.cell(row = row, column = 1, value = "Heat Sheet: ")
        row += 2

        for event in heat_sheet.dict_of_events:
            sheet.cell(row=row, column=1, value= "Event:")
            sheet.cell(row=row, column = 2, value=event)
            row += 1

            n_list = self.dict_of_events[event]
            for person in n_list:
                sheet.cell(row=row, column=1, value=person[1])
                sheet.cell(row=row, column=2, value=person[2])
                sheet.cell(row=row, column=3, value=person[3])
                sheet.cell(row=row, column=4, value=person[4])
                sheet.cell(row=row, column=5, value = "Heat: " + str(person[5]))
                sheet.cell(row=row, column=6, value = "Lane: " + str(person[6]))
                row += 1

            row += 1

            workbook.save(output)

        # with open(output, 'w') as out_file:
            # count = 0
            # out_file.write('Heat Sheet: ' + '\n\n')
            # for event in heat_sheet.dict_of_events:
                # out_file.write('Event: ' + event + '\n')
                # n_list = self.dict_of_events[event]
                # Sorts the list within each event by the lane number
                # n_list.sort(key=lambda x:x[6])
                # for person in n_list:
                    # out_file.write(
                        # person[1] + ' ' + person[2] + ' ' + person[3] + ' ' + person[4] + ' ' + 'Heat: ' + str(
                            # person[5]) + ' Lane: ' + str(person[6]) + '\n')
                # out_file.write('\n')


hs = HeatSheet()
hs.parser('example_simple_out.txt')
hs.create_heat_sheet('heat_sheet.xlsx')
